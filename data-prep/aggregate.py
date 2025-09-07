# CREATED: 21-JAN-2025
# LAST EDIT: 4-FEB-2025
# AUTHOR: DUANE RINEHART, MBA (drinehart@ucsd.edu)

'''TERMINAL APP FOR AGGREGATING RECORDINGS [PRIOR TO CREATING NWB FILE]'''

import os, pynwb
from datetime import datetime
import keyboard
import pyperclip
import json
from dotenv import load_dotenv
from pathlib import Path, PurePath
import polars as pl
import openpyxl
from openpyxl.styles import PatternFill
from prep import get_subject
import pytz
from tzlocal import get_localzone
import requests
import subprocess

#RECORDINGS STORED IN EXCEL FILES
#IF CHANGED, MODIFIY GLOBBING PATTERN BELOW
glob_pattern = ('.xlsx', '.xls')

dandi_api_url = "https://dandi.dk.ucsd.edu/api/"


# .env SHOULD BE STORED IN SAME DIRECTORY AS SCRIPT
script_dir = os.path.dirname(os.path.abspath(__file__))
env_path = os.path.join(script_dir, ".env")

# Check if the .env file exists
if not os.path.exists(env_path):
    # Create .env if it doesn't exist
    cwd = os.getcwd()
    output_path = Path(cwd, 'output')
    with open(env_path, "w") as env_file:
        env_file.write("EXP_MODALITY=1\n")
        env_file.write(f"INPUT_LOCATION='{cwd}'\n")
        env_file.write(f"OUTPUT_LOCATION='{output_path}'\n")
        env_file.write(f"EXPERIMENTERS='UNDEFINED'\n")
        env_file.write(f"INSTITUTION='UNDEFINED'\n")
        env_file.write(f"DANDI_API_KEY'\n")
load_dotenv()


def get_exp_common_name(exp_modality: str):
    if exp_modality == '1':
        exp_modality_name = 'EXTRACELLULAR ELECTROPHYSIOLOGY'
        summary_filename = 'input_ephys.xlsx'
    elif exp_modality == '2':
        exp_modality_name = 'WIDEFIELD IMAGING'
        summary_filename = 'input_widefield.xlsx'
    elif exp_modality == '3':
        exp_modality_name = '2PHOTON IMAGING'
        summary_filename = 'input_2photon.xlsx'
    elif exp_modality == '4':
        exp_modality_name = 'BEHAVIOR'
        summary_filename = 'input_behavior.xlsx'
    elif exp_modality == '5':
        exp_modality_name = 'fMRI'
        summary_filename = 'input_fMRI.xlsx'
    else:
        exp_modality_name = 'UNKNOWN'
        summary_filename = ''

    return exp_modality_name, summary_filename


def index_dir(input_path: str, glob_pattern: tuple[str, ...]) -> list[str]:
    """
    Indexes files in the specified directory that match the given glob pattern.

    Args:
        input_path (str): Path to the directory to search for files.
        glob_pattern (tuple[str, ...]): File extensions to look for (e.g., ('.xlsx', '.xls')).

    Returns:
        list[str]: A list of filenames matching the specified extensions.
    """

    # Ensure the path exists and is a directory
    if not os.path.exists(input_path) or not os.path.isdir(input_path):
        raise ValueError(f"The path '{input_path}' is not a valid directory.")

    # Find files that match the given extensions and exclude temporary files
    matching_files = [
        f for f in os.listdir(input_path) 
        if f.endswith(glob_pattern) and not f.startswith("~")
    ]

    return matching_files


def generate_experiments_summary():
    input_location = os.getenv("INPUT_LOCATION")
    meta_experimenters = os.getenv("EXPERIMENTERS")
    meta_institution = os.getenv("INSTITUTION")
    exp_modality_name, summary_filename = get_exp_common_name(os.getenv("EXP_MODALITY"))

    print(f'SUMMARY EXCEL FILE WILL BE STORED IN SAME FOLDER AS RECORDINGS: {input_location}')

    summary_file_path = Path(input_location, summary_filename)
    if os.path.exists(summary_file_path):
        print(f'FOUND PRE-EXISTING SUMMARY FILE: ({summary_file_path})')
    else:
        print(f'NO PRE-EXISTING SUMMARY FILE; CREATING BASED ON ../templates/{summary_filename}')
        source_file_path = Path("..", "templates", summary_filename)
        
        #CLEAN UP OLD EXCEL; ONLY NEED HEADERS
        df = pl.read_excel(source_file_path)
        df_first_row = df.head(1)
        # Build the filter expression using logical 'AND' across all columns
        filter_expr = pl.col(df_first_row.columns[0]).is_not_null()  # Start with the first column

        # Combine conditions for all columns using '&' (AND)
        for col in df_first_row.columns[1:]:
            filter_expr &= pl.col(col).is_not_null()

        # Apply the filter to keep only rows where all columns are non-null
        df_first_row_clean = df_first_row.filter(filter_expr)
        df_first_row_clean.write_excel(summary_file_path)

        wb = openpyxl.load_workbook(summary_file_path)
        ws = wb.active

        # Define the color fill (Orange)
        orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

        # Apply the orange fill to the first row
        for cell in ws[1]:
            cell.fill = orange_fill

        wb.save(summary_file_path)


    valid_files = index_dir(input_location, glob_pattern)
    print(f"FOUND {len(valid_files)} FILES IN DIRECTORY MATCHING PATTERN [{glob_pattern}]\n")

    data_rows = []
    for recording_file in valid_files:
        if not recording_file.startswith("recordings_"):
            continue

        print(f'PROCESSING {recording_file}...')
        
        #BUILD EACH ROW
        split_parts = recording_file.split('_')
        session_id = '_'.join(split_parts[4:-1])
        date_part, time_part = split_parts[2:4]

        date_obj = datetime.strptime(date_part, '%d.%b.%Y')
        time_obj = datetime.strptime(time_part, '%H.%M.%S')
        session_start_time = f"{date_obj.strftime('%Y-%m-%d')} {time_obj.strftime('%H:%M')}"

        subject_id = split_parts[4:5]
        print(subject_id)

        recordings_path = Path(input_location, recording_file)
        recording_dict = pl.read_excel(recordings_path).to_dict(as_series=False)

        json_str = json.dumps(recording_dict)
        
        #TODO - need to parse dict
                
        age_days = ''
        subject_description = ''
        genotype = ''
        sex = ''
        species = 'Mus musculus'
        weight = 'NA'
        subject_strain = 'NA'
        date_of_birth = 'Unkonwn'
        session_description = json_str
        recordings_folder_directory = 'NA'
        experimenters = meta_experimenters
        institution = meta_institution
        stimulus_notes_file	 = ''
        stimulus_notes_include = ''
        stimulus_notes_paradigm = ''
        surgery_notes = ''
        acquisition_mode = ''
        publication_figures = ''
        include_nwb = 'y'

        data = {
            "session_id": session_id,
            "session_start_time(YYYY-MM-DD HH:MM)": session_start_time,
            "subject_id": subject_id,
            "age(days)": age_days,
            "subject_description": subject_description,
            "genotype": genotype,
            "sex": sex,
            "species": species,
            "subject_weight": weight,
            "subject_strain": subject_strain,
            "date_of_birth(YYYY-MM-DD)": date_of_birth,
            "session_description": session_description,
            "recordings_folder_directory": recordings_folder_directory,
            "experimenters": experimenters,
            "institution": institution,
            "stimulus_notes_file": stimulus_notes_file,
            "stimulus_notes_include": stimulus_notes_include,
            "stimulus_notes_paradigm": stimulus_notes_paradigm,
            "surgery_notes": surgery_notes,
            "acquisition_mode": acquisition_mode,
            "publication_figures": publication_figures,
            "include_nwb": include_nwb
        }
        data_rows.append(data)

    new_data_df = pl.DataFrame(data_rows)

    # try:
    #     existing_df = pl.read_excel(summary_file_path)
    #     existing_columns = existing_df.columns
    #     new_data_df = new_data_df.select(existing_columns)

    #     # If new data is missing any columns, fill them with default values (e.g., None or NA)
    #     for col in existing_columns:
    #         if col not in new_data_df.columns:
    #             new_data_df = new_data_df.with_columns(pl.lit(None).alias(col))  # Add missing columns with None

    #     # Combine the existing DataFrame with the new data
    #     combined_df = existing_df.vstack(new_data_df)
        
    #     # Write the combined DataFrame back to the same Excel file
    #     combined_df.write_excel(summary_file_path)

    # except FileNotFoundError:
    #     print(f'ERROR WRITING TO FILE: {summary_file_path}')


def manage_meta_data():
    display_menu('meta_data')
    meta_experimenters = os.getenv("EXPERIMENTERS")
    meta_institution = os.getenv("INSTITUTION")
    meta_api_key = os.getenv("DANDI_API_KEY")

    while True:
        choice = input('SUBMENU - SELECTION: ')
        try:
            if int(choice) == 1: #EXPERIMENTERS
                meta_experimenters_new = input(f'PLEASE TYPE OR ENTER EXPERIMENTERS [ENTER for {meta_experimenters}]: ') or meta_experimenters_new

                if meta_experimenters_new is not meta_experimenters:
                    
                    #UPDATE .env FILE WITH SELECTION
                    with open(env_path, "r") as env_file:
                        lines = env_file.readlines()
                
                    with open(env_path, "w") as env_file:
                        found = False
                        for line in lines:
                            if line.startswith("EXPERIMENTERS="):
                                env_file.write(f"EXPERIMENTERS='{meta_experimenters_new}'\n")  # Replace with the new value
                                found = True
                            else:
                                env_file.write(line)
                        if not found:
                            env_file.write(f"EXPERIMENTERS='{meta_experimenters_new}'\n")  # Add if not present
                    
                    load_dotenv(override=True) #RELOAD ENVIRONMENT VARIABLES FROM .env FILE
                display_menu('meta_data')
                break
            elif int(choice) == 2: #INSTITUTION
                meta_institution_new = input(f'PLEASE TYPE OR ENTER INSTITUTION [ENTER for {meta_institution}]: ') or meta_institution_new
                if meta_institution_new is not meta_institution:

                    #UPDATE .env FILE WITH SELECTION
                    with open(env_path, "r") as env_file:
                        lines = env_file.readlines()
                
                    with open(env_path, "w") as env_file:
                        found = False
                        for line in lines:
                            if line.startswith("INSTITUTION="):
                                env_file.write(f"INSTITUTION={meta_experimenters_new}\n")  # Replace with the new value
                                found = True
                            else:
                                env_file.write(line)
                        if not found:
                            env_file.write(f"INSTITUTION='{meta_institution_new}'\n")  # Add if not present
                    
                    load_dotenv(override=True) #RELOAD ENVIRONMENT VARIABLES FROM .env FILE
                display_menu('meta_data')
                break
            elif int(choice) == 3: #api key
                print('YOU MUST LOGIN TO usarhythms.ucsd.edu TO RETRIEVE DANDI API KEY [UNDER USER SETTINGS]\n')
                meta_api_key_new = input(f'PLEASE TYPE OR ENTER DANDI API KEY [ENTER for {meta_api_key}]: ')
                if meta_api_key_new is not meta_api_key:
                    
                    #UPDATE .env FILE WITH SELECTION
                    with open(env_path, "r") as env_file:
                        lines = env_file.readlines()
                
                    with open(env_path, "w") as env_file:
                        found = False
                        for line in lines:
                            if line.startswith("DANDI_API_KEY="):
                                env_file.write(f"DANDI_API_KEY={meta_api_key_new.strip()}\n")
                                found = True
                            else:
                                env_file.write(line)
                        if not found:
                            env_file.write(f"DANDI_API_KEY={meta_api_key_new.strip()}\n")  # Add if not present

        except ValueError:
            if choice.upper() == 'M' or choice == '?':
                display_menu('meta_data')
            elif choice.upper() == 'R' or choice.upper() == 'X':
                display_menu()
                break
            else:
                print('INVALID SELECTION - (M)ENU')


def get_directory_path(input_location: str):
    """
    Used with manage_input_location() to define the source directory of recordings.
    
    Args:
        input_location (str): The default directory path to use if no input is provided.
    
    Returns:
        str: The absolute path of the selected directory, or None if the operation is canceled or the directory is invalid.
    """
    if not input_location:
        input_location = os.getcwd()
    
    print(f"PLEASE TYPE OR PASTE NEW DIRECTORY PATH [ENTER FOR '{input_location}', Ctrl+C to cancel]: ")
    
    try:
        user_input = input().strip()
        
        # If no input is provided, use the default location
        if not user_input:
            return os.path.abspath(input_location)
        
        # Check if the input is a valid directory
        if os.path.isdir(user_input):
            return os.path.abspath(user_input)
        else:
            print(f"Invalid directory: {user_input}")
            return None
    
    except KeyboardInterrupt:
        print("\nOperation canceled by the user.")
        return None


def manage_folder_locations():
    display_menu('folder_locations')
    input_location = os.getenv("INPUT_LOCATION")
    output_location = os.getenv("OUTPUT_LOCATION")

    while True:
        choice = input('SUBMENU - SELECTION: ')
        try:
            if int(choice) == 1: #DEFINE INPUT FOLDER
                absolute_path = get_directory_path(input_location)
                if absolute_path is None:
                    absolute_path = input_location

                # Check if the path is valid
                err = False
                if not os.path.exists(absolute_path):
                    print(f"ERROR: THE PATH '{absolute_path}' DOES NOT EXIST")
                    err = True
                else:
                    valid_files = index_dir(absolute_path, glob_pattern)
                    if not valid_files:
                        print("NO RECORDINGS FILES FOUND IN SPECIFIED DIRECTORY")
                        err = True
                if err == True:
                    proceed = input("ARE YOU SURE YOU WANT TO PROCEED? [yes/no]: ").strip().lower()
                    if proceed in ('no', 'n'):
                        break
                else:
                    print(f"FOUND {len(valid_files)} FILES IN DIRECTORY\n")
                    #UPDATE .env FILE WITH SELECTION
                    with open(env_path, "r") as env_file:
                        lines = env_file.readlines()
                
                    with open(env_path, "w") as env_file:
                        found = False
                        for line in lines:
                            if line.startswith("INPUT_LOCATION="):
                                env_file.write(f"INPUT_LOCATION={absolute_path}\n")  # Replace with the new value
                                found = True
                            else:
                                env_file.write(line)
                        if not found:
                            env_file.write(f"INPUT_LOCATION={absolute_path}\n")  # Add if not present
                    
                    load_dotenv(override=True) #RELOAD ENVIRONMENT VARIABLES FROM .env FILE
                    display_menu()
                    break
            elif int(choice) == 2: #SCAN INPUT FOLDER FOR RECORDINGS
                input_location = os.getenv("INPUT_LOCATION")
                valid_files = index_dir(input_location, glob_pattern)
                print(f"FOUND {len(valid_files)} FILES IN DIRECTORY MATCHING PATTERN [{glob_pattern}]\n")
            elif int(choice) == 3: #DEFINE OUTPUT FOLDER
                absolute_path = get_directory_path(output_location)
                if absolute_path is None:
                    absolute_path = output_location
                
                with open(env_path, "r") as env_file:
                    lines = env_file.readlines()

                with open(env_path, "w") as env_file:
                    found = False
                    for line in lines:
                        if line.startswith("OUTPUT_LOCATION="):
                            env_file.write(f"OUTPUT_LOCATION='{absolute_path}'\n")  # Replace with the new value
                            found = True
                        else:
                            env_file.write(line)
                    if not found:
                        env_file.write(f"OUTPUT_LOCATION={absolute_path}\n")  # Add if not present
                    
                load_dotenv(override=True) #RELOAD ENVIRONMENT VARIABLES FROM .env FILE
                display_menu()
                break
            else:
                print('INVALID SELECTION - (M)ENU')
        except ValueError:
            if choice.upper() == 'M' or choice == '?':
                display_menu('input_location')
            elif choice.upper() == 'R' or choice.upper() == 'X':
                display_menu()
                break
            else:
                print('INVALID SELECTION - (M)ENU')



def generate_nwb():
    ''' 1. READ SUMMARY EXCEL FOR RECORDINGS DATASETS TO INCLUDE (include_nwb = 'y')
        2. CREATE NWB SHELL WITH META-DATA (STORE IN TEMP FOLDER)
        3. VALIDATE / SHA256SUM
    '''
    input_location = os.getenv("INPUT_LOCATION")
    output_location = Path(os.getenv("OUTPUT_LOCATION"))
    output_location.mkdir(parents=True, exist_ok=True)
    
    exp_modality_name, summary_filename = get_exp_common_name(os.getenv("EXP_MODALITY"))
    summary_filename = Path(input_location, summary_filename)

    print(f'USING SUMMARY EXCEL FILE: {summary_filename}')
    print(f'NWB FILES WILL BE STAGED IN FOLDER: {output_location}\n')

    summary_df = pl.read_excel(summary_filename)
    
    
    summary_cols = summary_df.select([
        pl.col("session_id"),
        pl.col("session_start_time(YYYY-MM-DD HH:MM)")
            .str.strptime(pl.Datetime, format="%Y-%m-%d %H:%M")
            .map_elements(lambda x: x.replace(tzinfo=get_localzone())) # Add local timezone
            .alias("session_start_time_tz"),
        pl.col("subject_id"),
        pl.col("age(days)"),
        pl.col("subject_description"),
        pl.col("genotype"),
        pl.col("sex"),
        pl.col("species"),
        pl.col("subject_weight"),
        pl.col("subject_strain"),
        pl.col("date_of_birth(YYYY-MM-DD)").str.strptime(pl.Datetime, format="%Y-%m-%d %H:%M", strict=False),
        pl.col("session_description"),
        pl.col("recordings_folder_directory"),
        pl.col("experimenters"),
        pl.col("institution"),
        pl.col("stimulus_notes_file"),
        pl.col("stimulus_notes_include"),
        pl.col("stimulus_notes_paradigm"),
        pl.col("surgery_notes"),
        pl.col("acquisition_mode"),
        pl.col("publication_figures"),
        pl.col("include_nwb")
    ])

    # Loop through the rows and print the values
    for row in summary_cols.iter_rows():
        session_id, session_start_time, subject_id, age, subject_description, genotype, sex, species, subject_weight, subject_strain, date_of_birth, session_description, recordings_folder_directory, experimenters, institution, stimulus_notes_file, stimulus_notes_include, stimulus_notes_paradigm, surgery_notes, acquisition_mode, publication_figures, include_nwb = row
        if include_nwb == 'y':
            print(f"PROCESSING {session_id=} ...")
            recordings_file = Path(recordings_folder_directory, session_id)

            sex = 'U'  # unknown sex
            if sex is not None:
                if sex[0].upper() == 'M': #could me 'Male' or just 'M'
                    sex = 'M'
                elif sex[0].upper() == 'F' :
                    sex = 'F'

            print(f"{recordings_file=}")
            
            subject = get_subject(age, subject_description, genotype, sex, species, subject_id, subject_weight, date_of_birth, subject_strain)

            output_filename = Path(output_location, str(session_id) + '.nwb')

            # CONCATENATE STIMULUS NOTES (DEPENDS ON EXPERIMENT MODALITY)
            stimulus_notes = ''
            if os.getenv("EXP_MODALITY") == '4': #behavior
                print('NOT YET IMPLEMENTED')
                print(f'\tINCLUDING DATA FROM FILE: {stimulus_notes_file}')
            else:
                if stimulus_notes_include is not None and (stimulus_notes_include == 1 or stimulus_notes_include.lower() == 'y'):
                    stimulus_notes = "Stimulus paradigm: " + str(stimulus_notes_paradigm) + "; "

            #CREATE NWB FILE (BASIC META-DATA)
            nwbfile = pynwb.NWBFile(session_description = session_description,
                                    identifier = '',
                                    session_start_time = session_start_time,
                                    experiment_description = '',
                                    keywords = [str(experimenters), str(institution)],
                                    surgery = str(surgery_notes),
                                    pharmacology = '',
                                    stimulus_notes = str(stimulus_notes),
                                    experimenter = str(experimenters),
                                    institution = str(institution),
                                    subject = subject,
                                    notes = str(recordings_file)
                                    )

            #SAVE NWB FILE
            with pynwb.NWBHDF5IO(output_filename, 'w') as io:
                io.write(nwbfile)


def ingest_nwb():
    output_location = os.getenv("OUTPUT_LOCATION")
    meta_api_key = os.getenv("DANDI_API_KEY")
    if meta_api_key is None:
        raise ValueError('PLEASE SET DANDI API KEY IN .env FILE - OR IN META-DATA OPTION FROM MAIN MENU')
    elif output_location is None:
        print('PLEASE SET OUTPUT LOCATION IN .env FILE - OR IN META-DATA OPTION FROM MAIN MENU')
    else:
        os.environ["DANDI_API_KEY"] = meta_api_key
        output_location = Path(output_location)
        nwb_files = [f.name for f in output_location.iterdir() if f.is_file() and f.suffix == ".nwb"]
        print(f'INGESTING {len(nwb_files)} NWB FILES FROM \'{output_location}\'')

        dandi_url = f"{dandi_api_url}dandisets/"
        headers = {
            "Authorization": f"Token {meta_api_key}",  
            "Accept": "application/json",  # Optional: Specify the response format
        }

        for cnt, filename in enumerate(nwb_files):
            print(cnt, filename)

            nwb_file = Path(output_location, filename)
            io = pynwb.NWBHDF5IO(nwb_file, 'r')
            nwbfile_meta = io.read()

            payload = {
                "name": filename,
                "metadata": {
                    "contributor": [
                        {
                            "name": nwbfile_meta.experimenter[0],
                            "roleName": ["dcite:ContactPerson"]
                        }
                    ]
                }
            }

            print(payload)

            # Send the POST request to create the dandiset
            response = requests.post(dandi_url, headers=headers, json=payload, verify=True)

            if response.status_code == 201 or response.status_code == 200:
                # Get the dandiset ID from the response
                dandiset_id = response.json()["identifier"]
                print(f"Dandiset ID: {dandiset_id}")
                print(f"Response content: {response.text}")
            else:
                print(f"Request failed with status code: {response.status_code}")
                print(f"Response content: {response.text}")
            
            if cnt > 1:
                continue

def set_experiment_type():
    display_menu('exp_modality')

    while True:
        choice = input('SUBMENU - SELECTION: ')
        try:
            if int(choice) > 0 and int(choice) < 6:
                with open(env_path, "r") as env_file:
                    lines = env_file.readlines()
            
                with open(env_path, "w") as env_file:
                    found = False
                    for line in lines:
                        if line.startswith("EXP_MODALITY="):
                            env_file.write(f"EXP_MODALITY={choice}\n")  # Replace with the new value
                            found = True
                        else:
                            env_file.write(line)
                    if not found:
                        env_file.write(f"EXP_MODALITY={choice}\n")  # Add if not present
                
                load_dotenv(override=True) #RELOAD ENVIRONMENT VARIABLES FROM .env FILE
                display_menu()
                break
            else:
                print('INVALID SELECTION - (M)ENU')
        except ValueError:
            if choice.upper() == 'M' or choice == '?':
                display_menu('exp_modality')
            elif choice.upper() == 'R' or choice.upper() == 'X':
                display_menu()
                break
            else:
                print('INVALID SELECTION - (M)ENU')


def display_menu(options: str = 'main'):
    '''ENSURE CHOICE OPTIONS EXIST FOR EACH ITEM'''
    exp_modality_name, summary_filename = get_exp_common_name(os.getenv("EXP_MODALITY"))
    input_location = os.getenv("INPUT_LOCATION")
    output_location = os.getenv("OUTPUT_LOCATION")
    meta_experimenters = os.getenv("EXPERIMENTERS")
    meta_institution = os.getenv("INSTITUTION")
    meta_api_key = os.getenv("DANDI_API_KEY")

    if options == 'main':
        print(f'\n1 = SELECT EXPERIMENTAL MODALITY (CURRENT: {exp_modality_name})')
        print('2 = MANAGE RECORDINGS FOLDER LOCATIONS')
        print('3 = MANAGE GENERAL META-DATA (API-KEY, EXPERIMENTERS, INSTITUTION, ETC.)')
        print('4 = GENERATE SUMMARY EXCEL FILE')
        print('5 = CREATE NWB FILE')
        print('6 = INGEST NWB FILE TO U19 DANDI')
        print('M = DISPLAY MENU')
        print('X = EXIT\n')

    elif options == 'exp_modality':
        print(f'\nCURRENT SELECTION: {exp_modality_name}\n')
        print('1 = EXTRACELLULAR ELECTROPHYSIOLOGY')
        print('2 = WIDEFIELD IMAGING')
        print('3 = 2PHOTON IMAGING')
        print('4 = BEHAVIOR')
        print('5 = fMRI')
        print('M = DISPLAY MENU')
        print('R = RETURN TO MAIN MENU')

    elif options == 'folder_locations':
        print(f'\nCURRENT RECORDINGS INPUT LOCATION: \'{input_location}\'')
        print(f'\nCURRENT OUTPUT LOCATION: \'{output_location}\'\n')
        print('1 = ENTER NEW INPUT PATH')
        print('2 = SCAN INPUT PATH FOR RECORDINGS')
        print('3 = ENTER NEW OUTPUT PATH')
        print('M = DISPLAY MENU')
        print('R = RETURN TO MAIN MENU')

    elif options == 'meta_data':
        print(f'\nMANAGE META-DATA:\n')
        print(f'1 = SET EXPERIMENTER(S) [CURRENT: \'{meta_experimenters}\']')
        print(f'2 = SET INSTITUTION(S) [CURRENT: \'{meta_institution}\']')
        print(f'3 = SET DANDI API KEY [CURRENT: \'{meta_api_key}\']')
        print('M = DISPLAY MENU')
        print('R = RETURN TO MAIN MENU')


def main():
    print('\nWELCOME TO NWB AGGREGATOR 1.0')
    print('(c)2025 Kleinfeld Lab @ UCSD')
    print('------------------------------------')
    display_menu()

    while True:
        choice = input('MAIN MENU - ACTION: ')
        if choice == '1':
            set_experiment_type()
        elif choice == '2':
            manage_folder_locations()
        elif choice == '3':
            manage_meta_data()
        elif choice == '4':
            generate_experiments_summary()
        elif choice == '5':
            generate_nwb()
        elif choice == '6':
            ingest_nwb()
        elif choice.upper() == 'M':
            display_menu()
        elif choice.upper() == 'X':
            print('GOODBYE')
            break
        else:
            print('INVALID CHOICE - (M)ENU')


if __name__ == "__main__":
    main()