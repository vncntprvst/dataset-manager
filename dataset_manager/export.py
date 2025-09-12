from __future__ import annotations

import csv
import io
from typing import List, Any, Dict, Optional


def build_csv_bytes(columns: List[str], n_rows: int, rows: Optional[List[Dict[str, Any]]] = None) -> io.BytesIO:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(columns)
    if rows:
        for r in rows:
            writer.writerow([r.get(c, "") for c in columns])
    else:
        for _ in range(n_rows):
            writer.writerow([""] * len(columns))
    raw = buf.getvalue().encode("utf-8")
    return io.BytesIO(raw)


def build_workbook_bytes(columns: List[str], n_rows: int, rows: Optional[List[Dict[str, Any]]] = None) -> io.BytesIO:
    """Build XLSX workbook in-memory using pandas/openpyxl/xlsxwriter if available.

    Raises if no supported writer is available so caller can fall back to CSV.
    """
    # Try pandas with openpyxl/xlsxwriter engines
    try:
        import pandas as pd  # type: ignore

        if rows:
            # Build DataFrame from rows and ensure column order
            df = pd.DataFrame(rows)
            for c in columns:
                if c not in df.columns:
                    df[c] = None
            df = df[columns]
        else:
            df = pd.DataFrame(columns=columns, data=[[None] * len(columns) for _ in range(n_rows)])
        bio = io.BytesIO()
        # Prefer openpyxl; fall back to xlsxwriter
        used_engine = None
        try:
            df.to_excel(bio, index=False, engine="openpyxl")
            used_engine = "openpyxl"
        except Exception:
            df.to_excel(bio, index=False, engine="xlsxwriter")
            used_engine = "xlsxwriter"
        # Attempt header styling and freeze panes using openpyxl
        try:
            from openpyxl import load_workbook  # type: ignore
            from openpyxl.styles import PatternFill, Font  # type: ignore

            bio.seek(0)
            wb = load_workbook(bio)
            ws = wb.active
            # Freeze first row
            ws.freeze_panes = "A2"
            # Style header row (row 1)
            header_fill = PatternFill(fill_type="solid", fgColor="FFFFFF00")  # yellow
            header_font = Font(bold=True)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
            bio2 = io.BytesIO()
            wb.save(bio2)
            bio2.seek(0)
            return bio2
        except Exception:
            # If styling fails, return the raw workbook
            bio.seek(0)
            return bio
    except Exception as e:
        # Last-chance: minimal openpyxl without pandas
        try:
            from openpyxl import Workbook  # type: ignore

            wb = Workbook()
            ws = wb.active
            ws.title = "Template"
            # Header with style
            from openpyxl.styles import PatternFill, Font  # type: ignore
            header_fill = PatternFill(fill_type="solid", fgColor="FFFFFF00")  # yellow
            header_font = Font(bold=True)
            ws.append(columns)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
            if rows:
                for r in rows:
                    ws.append([r.get(c, None) for c in columns])
            else:
                for _ in range(n_rows):
                    ws.append([None] * len(columns))
            # Freeze top row
            ws.freeze_panes = "A2"
            bio = io.BytesIO()
            wb.save(bio)
            bio.seek(0)
            return bio
        except Exception as e2:
            # Propagate a meaningful combined error
            raise RuntimeError(
                "No Excel writer available (pandas/openpyxl/xlsxwriter missing)"
            ) from e2

